
from flask import Flask, render_template, request, jsonify, send_file
import random
import io
import csv
import traceback
from copy import deepcopy
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__)

# Configuration
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
TIME_SLOTS = [
    '09:00 - 10:00',
    '10:00 - 11:00',
    '11:00 - 12:00',
    '12:00 - 01:00',  # Lunch Option 1
    '01:00 - 02:00',  # Lunch Option 2
    '02:00 - 03:00',
    '03:00 - 04:00',
]
# Lunch slot options - classes can have lunch in different slots
LUNCH_OPTIONS = ['12:00 - 01:00', '01:00 - 02:00']


class GeneticTimetableGenerator:
    def __init__(self, lecturers, classes, constraints, class_info=None, 
                 population_size=50, generations=100, mutation_rate=0.15):
        self.lecturers = lecturers  # List of {name, hours_per_week, type}
        self.classes = classes
        self.constraints = constraints
        self.class_info = class_info or {}
        
        self.population_size = population_size
        self.generations = generations
        self.mutation_rate = mutation_rate

    def _is_teaching_cell(self, cell):
        """Return True when a cell represents an actual teaching slot."""
        if not cell:
            return False
        return cell.get('lecturer') not in ['LUNCH', 'Free', 'OFF']

    def _requires_break_after_two_consecutive(self):
        """Constraint flag: after 2 consecutive slots, lecturer must get 1 free slot."""
        return self.constraints.get('break_after_two_consecutive', True)

    def _requires_avoid_four_theory_continuous(self):
        """Constraint flag: avoid 4 continuous theory lectures in a day."""
        return self.constraints.get('avoid_four_theory_continuous', True)

    def _requires_labs_continuous(self):
        """Constraint flag: labs must be in 2 continuous slots."""
        return self.constraints.get('labs_require_two_continuous_slots', True)

    def _requires_avoid_unnecessary_gaps(self):
        """Constraint flag: avoid unnecessary internal gaps in a day."""
        return self.constraints.get('avoid_unnecessary_gaps', True)

    def _requires_workload_distribution(self):
        """Constraint flag: distribute lecturer workload across the week."""
        return self.constraints.get('distribute_workload_across_week', True)

    def _requires_time_of_day_balance(self):
        """Constraint flag: avoid assigning a lecturer in same day-part every day."""
        return self.constraints.get('balance_time_of_day_for_lecturers', True)

    def _violates_break_after_two_consecutive(self, schedule, day, slot_idx, duration, lecturer_name):
        """Check whether placing this lecturer would create 3 consecutive slots in a day."""
        day_lecturers = []
        for idx, slot in enumerate(TIME_SLOTS):
            if slot_idx <= idx < slot_idx + duration:
                day_lecturers.append(lecturer_name)
                continue

            cell = schedule[day][slot]
            day_lecturers.append(cell['lecturer'] if self._is_teaching_cell(cell) else None)

        for idx in range(len(day_lecturers) - 2):
            current = day_lecturers[idx]
            if current and current == day_lecturers[idx + 1] == day_lecturers[idx + 2]:
                return True
        return False

    def _count_break_after_two_consecutive_violations(self, schedule):
        """Count occurrences where a lecturer is assigned 3 consecutive slots in a day."""
        violations = 0
        for day in DAYS:
            day_lecturers = []
            for slot in TIME_SLOTS:
                cell = schedule[day][slot]
                day_lecturers.append(cell['lecturer'] if self._is_teaching_cell(cell) else None)

            for idx in range(len(day_lecturers) - 2):
                current = day_lecturers[idx]
                if current and current == day_lecturers[idx + 1] == day_lecturers[idx + 2]:
                    violations += 1
        return violations

    def _count_four_theory_continuous_violations(self, schedule):
        """Count sequences where 4 or more consecutive theory slots appear in a day."""
        violations = 0
        for day in DAYS:
            streak = 0
            for slot in TIME_SLOTS:
                cell = schedule[day][slot]
                is_theory = bool(cell) and cell.get('type') == 'theory' and self._is_teaching_cell(cell)
                if is_theory:
                    streak += 1
                else:
                    if streak >= 4:
                        violations += (streak - 3)
                    streak = 0
            if streak >= 4:
                violations += (streak - 3)
        return violations

    def _count_lab_continuity_violations(self, schedule):
        """Count lab/extra slots that are not paired with an adjacent same-lecturer slot."""
        violations = 0
        for day in DAYS:
            for idx, slot in enumerate(TIME_SLOTS):
                cell = schedule[day][slot]
                # Check for lab or extra-curricular that need continuous slots
                if not (cell and cell.get('type') in ['lab', 'extra'] and self._is_teaching_cell(cell)):
                    continue

                lecturer = cell.get('lecturer')
                cell_type = cell.get('type')
                has_left_pair = False
                has_right_pair = False

                if idx > 0:
                    left = schedule[day][TIME_SLOTS[idx - 1]]
                    has_left_pair = bool(left) and left.get('type') == cell_type and left.get('lecturer') == lecturer
                if idx < len(TIME_SLOTS) - 1:
                    right = schedule[day][TIME_SLOTS[idx + 1]]
                    has_right_pair = bool(right) and right.get('type') == cell_type and right.get('lecturer') == lecturer

                if not (has_left_pair or has_right_pair):
                    violations += 1
        return violations

    def _count_internal_gap_slots(self, schedule):
        """Count free slots that appear between two teaching slots on the same day."""
        gaps = 0
        for day in DAYS:
            teaching_indices = []
            for idx, slot in enumerate(TIME_SLOTS):
                if self._is_teaching_cell(schedule[day][slot]):
                    teaching_indices.append(idx)

            if len(teaching_indices) < 2:
                continue

            start = min(teaching_indices)
            end = max(teaching_indices)
            for idx in range(start, end + 1):
                cell = schedule[day][TIME_SLOTS[idx]]
                if cell and cell.get('lecturer') == 'Free':
                    gaps += 1
        return gaps

    def _count_workload_distribution_violations(self, schedule):
        """Count penalties for lecturers packed into too few days or overloaded single days."""
        penalties = 0
        for lecturer in self.lecturers:
            lecturer_name = lecturer['name']
            day_counts = []
            total = 0

            for day in DAYS:
                day_load = 0
                for slot in TIME_SLOTS:
                    cell = schedule[day][slot]
                    if self._is_teaching_cell(cell) and cell.get('lecturer') == lecturer_name:
                        day_load += 1
                day_counts.append(day_load)
                total += day_load

            if total == 0:
                continue

            days_used = sum(1 for d in day_counts if d > 0)
            min_days_expected = min(4, max(2, (total + 1) // 2)) if total >= 2 else 1
            if days_used < min_days_expected:
                penalties += (min_days_expected - days_used)

            max_day_load = max(day_counts)
            if total >= 4 and max_day_load > 2:
                penalties += (max_day_load - 2)

        return penalties

    def _count_time_of_day_balance_violations(self, schedule):
        """Penalize lecturers repeatedly fixed to morning or first afternoon slot across days."""
        penalties = 0
        morning_indices = {0, 1, 2}  # 09:00-12:00
        first_afternoon_idx = 4      # 01:00-02:00

        for lecturer in self.lecturers:
            lecturer_name = lecturer['name']
            teaching_days = 0
            morning_days = 0
            first_afternoon_days = 0

            for day in DAYS:
                teaches_today = False
                has_morning = False
                has_first_afternoon = False

                for idx, slot in enumerate(TIME_SLOTS):
                    cell = schedule[day][slot]
                    if not (self._is_teaching_cell(cell) and cell.get('lecturer') == lecturer_name):
                        continue

                    teaches_today = True
                    if idx in morning_indices:
                        has_morning = True
                    if idx == first_afternoon_idx:
                        has_first_afternoon = True

                if teaches_today:
                    teaching_days += 1
                    if has_morning:
                        morning_days += 1
                    if has_first_afternoon:
                        first_afternoon_days += 1

            # Apply only when a lecturer is spread across multiple days.
            if teaching_days < 3:
                continue

            if morning_days == teaching_days:
                penalties += (teaching_days - 2)
            if first_afternoon_days == teaching_days:
                penalties += (teaching_days - 2)

            # Soft penalty for heavy day-part bias even if not absolute.
            if morning_days / teaching_days > 0.8:
                penalties += 1
            if first_afternoon_days / teaching_days > 0.8:
                penalties += 1

        return penalties

    def _is_mutable_slot(self, day, slot_idx, lunch_slot):
        """Mutation can only swap real working slots, not lunch or off periods."""
        slot = TIME_SLOTS[slot_idx]
        if day == 'Saturday' and slot_idx >= 4:
            return False
        if day != 'Saturday' and slot == lunch_slot:
            return False
        return True
    
    def get_lunch_slot(self, cls):
        """Get lunch slot for a specific class"""
        if cls in self.class_info and self.class_info[cls].get('lunch_slot'):
            return self.class_info[cls]['lunch_slot']
        class_idx = self.classes.index(cls) if cls in self.classes else 0
        return LUNCH_OPTIONS[class_idx % len(LUNCH_OPTIONS)]
    
    def create_individual(self):
        """Create a random timetable (individual/chromosome)"""
        individual = {}
        for cls in self.classes:
            individual[cls] = self._random_schedule(cls)
        return individual
    
    def _random_schedule(self, cls):
        """Create random schedule for a class with lecturers placed randomly"""
        schedule = {day: {slot: None for slot in TIME_SLOTS} for day in DAYS}
        
        # Add lunch break (class-specific) - ONLY on Monday-Friday, NOT Saturday
        lunch_slot = self.get_lunch_slot(cls)
        if lunch_slot not in LUNCH_OPTIONS:
            lunch_slot = LUNCH_OPTIONS[0]
        
        # Add lunch break ONLY for Monday-Friday
        for day in DAYS:
            if day != 'Saturday':
                schedule[day][lunch_slot] = {'lecturer': 'LUNCH', 'type': 'break'}
        
        # Shuffle lecturers for randomness
        lecturers_copy = deepcopy(self.lecturers)
        random.shuffle(lecturers_copy)
        
        # Place each lecturer's hours
        for lecturer in lecturers_copy:
            lecturer_name = lecturer['name']
            hours_needed = lecturer['hours_per_week']
            # Labs and extra-curricular with 2+ hours need 2 continuous slots
            lecturer_type = lecturer.get('type', 'theory')
            duration = 2 if lecturer_type in ['lab', 'extra'] and hours_needed >= 2 else 1
            hours_placed = 0
            
            # Try to place lecturer hours
            attempts = 0
            max_attempts = 100
            
            while hours_placed < hours_needed and attempts < max_attempts:
                day = random.choice(DAYS)
                
                # For Saturday, only use first 4 slots (09:00-13:00)
                if day == 'Saturday':
                    slot_idx = random.randint(0, min(3, len(TIME_SLOTS) - duration))
                else:
                    slot_idx = random.randint(0, len(TIME_SLOTS) - duration)
                
                # Check if we can place
                can_place = True
                for i in range(duration):
                    slot = TIME_SLOTS[slot_idx + i]
                    
                    # No lunch slot check for Saturday
                    if day != 'Saturday' and slot == lunch_slot:
                        can_place = False
                        break
                    
                    # Saturday: only first 4 slots
                    if day == 'Saturday' and slot_idx + i >= 4:
                        can_place = False
                        break
                    
                    if schedule[day][slot] is not None:
                        can_place = False
                        break
                
                if can_place:
                    if self._requires_break_after_two_consecutive() and self._violates_break_after_two_consecutive(
                        schedule, day, slot_idx, duration, lecturer_name
                    ):
                        can_place = False

                if can_place:
                    # Place the lecturer
                    for i in range(duration):
                        slot = TIME_SLOTS[slot_idx + i]
                        schedule[day][slot] = {'lecturer': lecturer_name, 'type': lecturer.get('type', 'theory')}
                    hours_placed += duration
                
                attempts += 1
        
        # Fill remaining free periods
        for day in DAYS:
            if day != 'Saturday':
                # Monday-Friday: fill all empty slots except lunch
                for slot in TIME_SLOTS:
                    if slot != lunch_slot and schedule[day][slot] is None:
                        schedule[day][slot] = {'lecturer': 'Free', 'type': 'free'}
            else:
                # Saturday: only 4 hours (09:00-13:00), rest is off
                for idx, slot in enumerate(TIME_SLOTS):
                    if idx < 4:  # First 4 slots (09:00-13:00)
                        if schedule[day][slot] is None:
                            schedule[day][slot] = {'lecturer': 'Free', 'type': 'free'}
                    else:  # After 13:00
                        schedule[day][slot] = {'lecturer': 'OFF', 'type': 'off'}
        
        return schedule
    
    def calculate_fitness(self, individual):
        """Calculate fitness: how good the solution is. Higher = better"""
        fitness = 0
        penalty = 0
        
        for cls in self.classes:
            schedule = individual[cls]
            
            # 1. Count lecturers scheduled
            lecturers_placed = {}
            for day in DAYS:
                for slot in TIME_SLOTS:
                    cell = schedule[day][slot]
                    if cell and cell.get('lecturer') not in ['LUNCH', 'Free']:
                        lecturer = cell['lecturer']
                        lecturers_placed[lecturer] = lecturers_placed.get(lecturer, 0) + 1
            
            # 2. Bonus for meeting hour requirements for each lecturer
            for lecturer in self.lecturers:
                lecturer_name = lecturer['name']
                if lecturer_name in lecturers_placed:
                    hours_placed = lecturers_placed[lecturer_name]
                    hours_needed = lecturer['hours_per_week']
                    if hours_placed >= hours_needed:
                        fitness += 20  # Bonus for meeting requirement
                    else:
                        penalty += (hours_needed - hours_placed) * 3  # Penalty for not meeting
            
            # 3. Bonus for balanced distribution (spread across days)
            days_used = sum(1 for day in DAYS if any(
                schedule[day][slot] and 
                schedule[day][slot].get('lecturer') not in ['LUNCH', 'Free']
                for slot in TIME_SLOTS
            ))
            fitness += days_used * 5
            
            # 4. Penalty for too many free periods
            free_periods = sum(1 for day in DAYS 
                             for slot in TIME_SLOTS
                             if schedule[day][slot] and schedule[day][slot].get('lecturer') == 'Free')
            penalty += free_periods

            # 5. Penalty when a lecturer teaches 3 consecutive slots without a break
            if self._requires_break_after_two_consecutive():
                violations = self._count_break_after_two_consecutive_violations(schedule)
                penalty += violations * 15

            # 6. Penalty for 4+ continuous theory lectures
            if self._requires_avoid_four_theory_continuous():
                violations = self._count_four_theory_continuous_violations(schedule)
                penalty += violations * 12

            # 7. Penalty if lab slots are not contiguous pairs
            if self._requires_labs_continuous():
                violations = self._count_lab_continuity_violations(schedule)
                penalty += violations * 20

            # 8. Penalty for unnecessary internal gaps
            if self._requires_avoid_unnecessary_gaps():
                gaps = self._count_internal_gap_slots(schedule)
                penalty += gaps * 5

            # 9. Penalty for uneven lecturer workload distribution
            if self._requires_workload_distribution():
                violations = self._count_workload_distribution_violations(schedule)
                penalty += violations * 8

            # 10. Penalty for repetitive day-part allocation per lecturer
            if self._requires_time_of_day_balance():
                violations = self._count_time_of_day_balance_violations(schedule)
                penalty += violations * 10
        
        return fitness - penalty
    
    def crossover(self, parent1, parent2):
        """Combine two solutions to create offspring"""
        child = {}
        for cls in self.classes:
            if random.random() < 0.5:
                child[cls] = deepcopy(parent1[cls])
            else:
                child[cls] = deepcopy(parent2[cls])
        return child
    
    def mutate(self, individual):
        """Randomly modify the timetable"""
        mutated = deepcopy(individual)
        
        for cls in self.classes:
            if random.random() < self.mutation_rate:
                schedule = mutated[cls]
                lunch_slot = self.get_lunch_slot(cls)

                # Random mutation: swap two valid working slots only.
                candidates = []
                for day in DAYS:
                    for slot_idx, slot in enumerate(TIME_SLOTS):
                        if self._is_mutable_slot(day, slot_idx, lunch_slot):
                            candidates.append((day, slot))

                if len(candidates) >= 2:
                    (day1, slot1), (day2, slot2) = random.sample(candidates, k=2)

                    # Never overwrite fixed non-working cells.
                    if schedule[day1][slot1].get('lecturer') not in ['LUNCH', 'OFF'] and \
                       schedule[day2][slot2].get('lecturer') not in ['LUNCH', 'OFF']:
                        schedule[day1][slot1], schedule[day2][slot2] = \
                            schedule[day2][slot2], schedule[day1][slot1]
        
        return mutated
    
    def generate(self):
        """Run genetic algorithm to generate timetable"""
        print(f"Starting Genetic Algorithm: population={self.population_size}, generations={self.generations}")
        
        with open('debug.log', 'a', encoding='utf-8') as f:
            f.write(f"\n=== GA STARTING ===\n")
            f.write(f"Population: {self.population_size}, Generations: {self.generations}\n")
        
        # Step 1: Create initial population
        population = [self.create_individual() for _ in range(self.population_size)]
        
        best_individual = None
        best_fitness = -float('inf')
        
        # Step 2: Evolution loop
        for generation in range(self.generations):
            # Calculate fitness for all individuals
            fitness_scores = [self.calculate_fitness(ind) for ind in population]
            
            # Track best solution
            max_fitness = max(fitness_scores)
            max_idx = fitness_scores.index(max_fitness)
            
            if max_fitness > best_fitness:
                best_fitness = max_fitness
                best_individual = deepcopy(population[max_idx])
            
            if generation % 10 == 0:
                print(f"Generation {generation}: Best Fitness = {best_fitness:.2f}")
                with open('debug.log', 'a', encoding='utf-8') as f:
                    f.write(f"Generation {generation}: Fitness = {best_fitness:.2f}\n")
            
            # Step 3: Selection (tournament selection)
            new_population = []
            for _ in range(self.population_size):
                # Select 3 random individuals, take the best
                tournament_size = min(3, self.population_size)
                tournament_indices = random.sample(range(self.population_size), tournament_size)
                
                # Find best in tournament
                best_tournament_idx = tournament_indices[0]
                best_tournament_fitness = fitness_scores[best_tournament_idx]
                
                for idx in tournament_indices[1:]:
                    if fitness_scores[idx] > best_tournament_fitness:
                        best_tournament_fitness = fitness_scores[idx]
                        best_tournament_idx = idx
                
                new_population.append(deepcopy(population[best_tournament_idx]))
            
            # Step 4: Crossover and mutation combined
            next_population = []
            for i in range(0, len(new_population), 2):
                parent1 = new_population[i]
                parent2 = new_population[i + 1] if i + 1 < len(new_population) else new_population[0]
                
                # Create two children
                child1 = self.crossover(parent1, parent2)
                child2 = self.crossover(parent2, parent1)
                
                # Apply mutation
                child1 = self.mutate(child1)
                child2 = self.mutate(child2)
                
                next_population.extend([child1, child2])
            
            # Trim to population size
            population = next_population[:self.population_size]
        
        print(f"\nBest Solution Found - Fitness: {best_fitness:.2f}")
        with open('debug.log', 'a', encoding='utf-8') as f:
            f.write(f"GA Complete - Best Fitness: {best_fitness:.2f}\n")
        
        return self.format_timetable(best_individual)
    
    def format_timetable(self, individual):
        """Format timetable for JSON response"""
        formatted = {}
        for cls in self.classes:
            formatted[cls] = {}
            for day in DAYS:
                formatted[cls][day] = {}
                for slot in TIME_SLOTS:
                    cell = individual[cls][day][slot]
                    if cell is None:
                        formatted[cls][day][slot] = {'lecturer': 'Free', 'type': 'free'}
                    else:
                        formatted[cls][day][slot] = cell
        return formatted


@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')


@app.route('/generate', methods=['POST'])
def generate():
    """Generate timetable based on input data"""
    try:
        data = request.json
        
        # Handle both 'lecturers' (new wizard) and 'subjects' (old UI) keys
        lecturers = data.get('lecturers', data.get('subjects', []))
        classes_data = data.get('classes', [])
        constraints = data.get('constraints', {})
        
        if not lecturers or not classes_data:
            return jsonify({'error': 'Please provide lecturers and classes'}), 400
        
        # Handle classes - could be strings or objects
        class_names = []
        class_info = {}
        for cls in classes_data:
            if isinstance(cls, str):
                class_names.append(cls)
            elif isinstance(cls, dict):
                name = cls.get('name', '')
                if name:
                    class_names.append(name)
                    class_info[name] = {
                        'lunch_slot': cls.get('lunch_slot', ''),
                        'size': cls.get('students', cls.get('size', 60))
                    }
        
        # Use Genetic Algorithm
        generator = GeneticTimetableGenerator(
            lecturers, 
            class_names, 
            constraints, 
            class_info,
            population_size=50,
            generations=100,
            mutation_rate=0.15
        )
        timetable = generator.generate()
        
        return jsonify({
            'success': True,
            'timetable': timetable,
            'days': DAYS,
            'timeSlots': TIME_SLOTS,
            'classes': class_names,
            'lunchOptions': LUNCH_OPTIONS
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/download', methods=['POST'])
def download():
    """Download timetable as CSV"""
    try:
        data = request.json
        timetable = data.get('timetable', {})
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        for cls, schedule in timetable.items():
            writer.writerow([f'=== {cls} ==='])
            writer.writerow(['Time'] + DAYS)
            
            for slot in TIME_SLOTS:
                row = [slot]
                for day in DAYS:
                    cell = schedule.get(day, {}).get(slot, {})
                    if isinstance(cell, dict):
                        text = cell.get('lecturer', '')
                        if cell.get('type'):
                            text += f" ({cell['type']})"
                    else:
                        text = str(cell) if cell else ''
                    row.append(text)
                writer.writerow(row)
            
            writer.writerow([])
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='timetable.csv'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    """Upload Excel file with subjects, classes, and rooms data"""
    try:
        # Write debug to file
        with open('debug.log', 'a') as f:
            f.write("\n" + "=" * 80 + "\n")
            f.write("UPLOAD EXCEL CALLED\n")
            f.write("=" * 80 + "\n")
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        with open('debug.log', 'a') as f:
            f.write(f"File name: {file.filename}\n")
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400
        
        # Read Excel file
        with open('debug.log', 'a') as f:
            f.write("Reading Excel file...\n")
        
        excel_data = pd.ExcelFile(file)
        
        with open('debug.log', 'a') as f:
            f.write(f"Sheet names: {excel_data.sheet_names}\n")
        
        subjects = []
        classes = []
        rooms = []
        
        # Normalize sheet names (remove trailing spaces)
        sheet_names_normalized = {name.strip(): name for name in excel_data.sheet_names}
        
        # Read Subjects sheet
        subjects_sheet = sheet_names_normalized.get('Subjects')
        if subjects_sheet:
            try:
                df_subjects = pd.read_excel(excel_data, sheet_name=subjects_sheet, header=0)
                df_subjects = df_subjects.fillna('')
                
                with open('debug.log', 'a') as f:
                    f.write(f"Columns: {list(df_subjects.columns)}\n")
                    f.write(f"Shape: {df_subjects.shape}\n")
                    f.write(f"Data:\n{df_subjects.to_string()}\n")
                
                # Try using iloc to access columns by position instead of name
                for idx, row in df_subjects.iterrows():
                    # Column 0 = Subject Name, 1 = Type, 2 = Hours/Week, 3 = Teacher, 4 = For Classes
                    try:
                        subject_name = str(row.iloc[0]).strip() if len(row) > 0 else ''
                        subject_type = str(row.iloc[1]).strip().lower() if len(row) > 1 else 'theory'
                        hours = int(float(str(row.iloc[2]))) if len(row) > 2 else 3
                        teacher = str(row.iloc[3]).strip() if len(row) > 3 else ''
                        for_classes_str = str(row.iloc[4]).strip() if len(row) > 4 else ''
                        
                        with open('debug.log', 'a') as f:
                            f.write(f"Row {idx}: name={subject_name}, type={subject_type}, hours={hours}, teacher={teacher}, classes={for_classes_str}\n")
                        
                        if subject_name and subject_name.lower() != 'subject name':
                            for_classes = []
                            if for_classes_str and for_classes_str.lower() != 'all':
                                # Split by comma and clean up each class name
                                raw_classes = [c.strip() for c in for_classes_str.split(',') if c.strip()]
                                for rc in raw_classes:
                                    # Normalize: remove quotes and extra spaces
                                    normalized = rc.replace("'", "").replace('"', '').strip()
                                    if normalized:
                                        for_classes.append(normalized)
                            
                            subjects.append({
                                'name': subject_name,
                                'type': subject_type,
                                'hours_per_week': hours,
                                'teacher': teacher,
                                'for_classes': for_classes,
                                'combined_classes': []
                            })
                            with open('debug.log', 'a', encoding='utf-8') as f:
                                f.write(f"Added subject: {subject_name} -> for_classes list: {for_classes}\n")
                    except Exception as e:
                        with open('debug.log', 'a') as f:
                            f.write(f"Error processing row {idx}: {e}\n")
                
                with open('debug.log', 'a') as f:
                    f.write(f"Total subjects loaded: {len(subjects)}\n")
            except Exception as e:
                with open('debug.log', 'a') as f:
                    f.write(f"Error reading Subjects sheet: {e}\n")
                    f.write(traceback.format_exc())
        
        # Read Classes sheet
        classes_sheet = sheet_names_normalized.get('Classes')
        if classes_sheet:
            df_classes = pd.read_excel(excel_data, sheet_name=classes_sheet)
            df_classes = df_classes.fillna('')
            # Normalize column names
            df_classes.columns = df_classes.columns.str.strip().str.lower()
            
            for _, row in df_classes.iterrows():
                # Get class name
                class_name = None
                for col in ['class name', 'class', 'name']:
                    try:
                        if col in df_classes.columns and str(row[col]).strip():
                            class_name = str(row[col]).strip()
                            break
                    except:
                        pass
                
                if class_name:
                    # Get students
                    students = 60
                    for col in ['students', 'student count', 'size']:
                        try:
                            if col in df_classes.columns:
                                students = int(float(str(row[col])))
                                break
                        except:
                            pass
                    
                    # Get lunch slot
                    lunch_slot = ''
                    for col in ['lunch slot', 'lunch_slot']:
                        try:
                            if col in df_classes.columns:
                                lunch_slot = str(row[col]).strip()
                                break
                        except:
                            pass
                    
                    # Validate lunch slot
                    if lunch_slot and lunch_slot not in LUNCH_OPTIONS:
                        lunch_slot = ''  # Reset to default if invalid
                    
                    classes.append({
                        'name': class_name,
                        'size': students,
                        'lunch_slot': lunch_slot
                    })
        
        with open('debug.log', 'a') as f:
            f.write(f"\nClasses loaded: {[c['name'] for c in classes]}\n")
        
        # Read Rooms sheet
        rooms_sheet = sheet_names_normalized.get('Rooms')
        if rooms_sheet:
            df_rooms = pd.read_excel(excel_data, sheet_name=rooms_sheet)
            df_rooms = df_rooms.fillna('')
            # Normalize column names
            df_rooms.columns = df_rooms.columns.str.strip().str.lower()
            
            for _, row in df_rooms.iterrows():
                # Get room name
                room_name = None
                for col in ['room name', 'room', 'name']:
                    try:
                        if col in df_rooms.columns and str(row[col]).strip():
                            room_name = str(row[col]).strip()
                            break
                    except:
                        pass
                
                if room_name:
                    # Get type
                    room_type = 'theory'
                    for col in ['type', 'room type']:
                        try:
                            if col in df_rooms.columns:
                                room_type = str(row[col]).strip().lower()
                                if room_type:
                                    break
                        except:
                            pass
                    
                    # Get capacity
                    capacity = 60
                    for col in ['capacity', 'cap']:
                        try:
                            if col in df_rooms.columns:
                                capacity = int(float(str(row[col])))
                                break
                        except:
                            pass
                    
                    rooms.append({
                        'name': room_name,
                        'type': room_type,
                        'capacity': capacity
                    })
        
        return jsonify({
            'success': True,
            'subjects': subjects,
            'classes': classes,
            'rooms': rooms
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/download-template')
def download_template():
    """Download Excel template for input data"""
    try:
        wb = Workbook()
        
        # Styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Subjects Sheet
        ws_subjects = wb.active
        ws_subjects.title = 'Subjects'
        headers = ['Subject Name', 'Type', 'Hours/Week', 'Teacher', 'For Classes']
        for col, header in enumerate(headers, 1):
            cell = ws_subjects.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data for Subjects - showing different subjects for different classes
        sample_subjects = [
            ['Mathematics', 'theory', 4, 'Dr. Smith', 'All'],
            ['English', 'theory', 3, 'Mrs. Davis', 'All'],
            ['Physics', 'theory', 3, 'Dr. Johnson', 'CSE Sem-1, ECE Sem-1'],
            ['Physics Lab', 'lab', 2, 'Dr. Johnson', 'CSE Sem-1, ECE Sem-1'],
            ['Data Structures', 'theory', 4, 'Prof. Kumar', 'CSE Sem-3'],
            ['DSA Lab', 'lab', 2, 'Prof. Kumar', 'CSE Sem-3'],
            ['Digital Electronics', 'theory', 3, 'Dr. Sharma', 'ECE Sem-1'],
            ['DE Lab', 'lab', 2, 'Dr. Sharma', 'ECE Sem-1'],
            ['Programming in C', 'theory', 3, 'Mr. Brown', 'CSE Sem-1'],
            ['C Programming Lab', 'lab', 2, 'Mr. Brown', 'CSE Sem-1'],
            ['DBMS', 'theory', 3, 'Prof. Williams', 'CSE Sem-3'],
            ['DBMS Lab', 'lab', 2, 'Prof. Williams', 'CSE Sem-3'],
        ]
        for row_idx, row_data in enumerate(sample_subjects, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_subjects.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Adjust column widths
        ws_subjects.column_dimensions['A'].width = 20
        ws_subjects.column_dimensions['B'].width = 12
        ws_subjects.column_dimensions['C'].width = 12
        ws_subjects.column_dimensions['D'].width = 20
        ws_subjects.column_dimensions['E'].width = 25
        
        # Classes Sheet
        ws_classes = wb.create_sheet('Classes')
        headers = ['Class Name', 'Students', 'Lunch Slot']
        for col, header in enumerate(headers, 1):
            cell = ws_classes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        sample_classes = [
            ['CSE Sem-1', 60, '12:00 - 01:00'],
            ['CSE Sem-3', 55, '01:00 - 02:00'],
            ['ECE Sem-1', 50, '12:00 - 01:00'],
        ]
        for row_idx, row_data in enumerate(sample_classes, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_classes.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        ws_classes.column_dimensions['A'].width = 20
        ws_classes.column_dimensions['B'].width = 12
        ws_classes.column_dimensions['C'].width = 18
        
        # Rooms Sheet
        ws_rooms = wb.create_sheet('Rooms')
        headers = ['Room Name', 'Type', 'Capacity']
        for col, header in enumerate(headers, 1):
            cell = ws_rooms.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        sample_rooms = [
            ['Room 101', 'theory', 60],
            ['Room 102', 'theory', 60],
            ['Room 103', 'theory', 40],
            ['Computer Lab', 'lab', 30],
            ['Physics Lab', 'lab', 30],
            ['Chemistry Lab', 'lab', 30],
        ]
        for row_idx, row_data in enumerate(sample_rooms, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_rooms.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        ws_rooms.column_dimensions['A'].width = 18
        ws_rooms.column_dimensions['B'].width = 12
        ws_rooms.column_dimensions['C'].width = 12
        
        # Instructions Sheet
        ws_instructions = wb.create_sheet('Instructions')
        instructions = [
            ['TIMETABLE INPUT TEMPLATE - INSTRUCTIONS'],
            [''],
            ['This Excel file contains 3 sheets for input:'],
            [''],
            ['1. SUBJECTS SHEET:'],
            ['   - Subject Name: Name of the subject (required)'],
            ['   - Type: Either "theory" or "lab" (labs are 2-hour sessions)'],
            ['   - Hours/Week: Number of hours per week for this subject'],
            ['   - Teacher: Name of the teacher'],
            ['   - For Classes: Which classes this subject is for'],
            ['       * Use "All" or leave empty to assign to all classes'],
            ['       * Use comma-separated class names for specific classes'],
            ['       * Example: "CSE Sem-1, CSE Sem-3" or "ECE Sem-1"'],
            [''],
            ['2. CLASSES SHEET:'],
            ['   - Class Name: Name/ID of the class (e.g., "CSE Sem-1")'],
            ['   - Students: Number of students in the class'],
            ['   - Lunch Slot: When this class has lunch break'],
            ['       * Options: "12:00 - 01:00" or "01:00 - 02:00"'],
            ['       * Different classes can have different lunch times'],
            ['       * Leave empty for auto-assignment'],
            [''],
            ['3. ROOMS SHEET:'],
            ['   - Room Name: Name/ID of the room'],
            ['   - Type: Either "theory" or "lab"'],
            ['   - Capacity: Maximum number of students'],
            [''],
            ['IMPORTANT: Class names in "For Classes" column must match'],
            ['exactly with names in the Classes sheet.'],
            [''],
            ['NOTE: Sample data is provided. Replace with your actual data.'],
        ]
        for row_idx, row_data in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row_idx, column=1, value=row_data[0])
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
        ws_instructions.column_dimensions['A'].width = 60
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='timetable_input_template.xlsx'
        )
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/download-excel', methods=['POST'])
def download_excel():
    """Download generated timetable as Excel"""
    try:
        data = request.json
        timetable = data.get('timetable', {})
        classes_list = data.get('classes', [])
        
        wb = Workbook()
        
        # Styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        day_fill = PatternFill(start_color='cce5ff', end_color='cce5ff', fill_type='solid')
        lunch_fill = PatternFill(start_color='ffd700', end_color='ffd700', fill_type='solid')
        lab_fill = PatternFill(start_color='ffe6e6', end_color='ffe6e6', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        first_sheet = True
        for cls in classes_list:
            if first_sheet:
                ws = wb.active
                ws.title = cls[:31]  # Excel sheet name limit
                first_sheet = False
            else:
                ws = wb.create_sheet(cls[:31])
            
            schedule = timetable.get(cls, {})
            
            # Title
            ws.merge_cells('A1:I1')
            title_cell = ws.cell(row=1, column=1, value=f'TIME TABLE - {cls}')
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Headers (Time slots)
            ws.cell(row=3, column=1, value='DAY').font = Font(bold=True)
            ws.cell(row=3, column=1).fill = day_fill
            ws.cell(row=3, column=1).border = thin_border
            ws.cell(row=3, column=1).alignment = center_align
            
            for col_idx, slot in enumerate(TIME_SLOTS, 2):
                cell = ws.cell(row=3, column=col_idx, value=slot)
                cell.font = Font(bold=True, size=9)
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            
            # Days and schedule
            for row_idx, day in enumerate(DAYS, 4):
                day_cell = ws.cell(row=row_idx, column=1, value=day.upper())
                day_cell.font = Font(bold=True)
                day_cell.fill = day_fill
                day_cell.border = thin_border
                day_cell.alignment = center_align
                
                for col_idx, slot in enumerate(TIME_SLOTS, 2):
                    cell_data = schedule.get(day, {}).get(slot, {})
                    
                    if cell_data.get('type') == 'break':
                        content = 'LUNCH BREAK'
                        cell = ws.cell(row=row_idx, column=col_idx, value=content)
                        cell.fill = lunch_fill
                    elif cell_data.get('type') == 'lab':
                        content = f"{cell_data.get('lecturer', '')}\n({cell_data.get('type', '').upper()})"
                        cell = ws.cell(row=row_idx, column=col_idx, value=content)
                        cell.fill = lab_fill
                    elif cell_data.get('lecturer') and cell_data.get('lecturer') not in ['Free', 'OFF']:
                        content = f"{cell_data.get('lecturer', '')}\n({cell_data.get('type', 'Theory')})"
                        cell = ws.cell(row=row_idx, column=col_idx, value=content)
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx, value='-')
                    
                    cell.border = thin_border
                    cell.alignment = center_align
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            for col_idx in range(2, len(TIME_SLOTS) + 2):
                ws.column_dimensions[chr(64 + col_idx)].width = 14
            
            # Row heights
            for row_idx in range(4, 4 + len(DAYS)):
                ws.row_dimensions[row_idx].height = 40
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='generated_timetable.xlsx'
        )
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
